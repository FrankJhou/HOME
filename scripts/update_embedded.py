#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
石化品儀表板 · 內嵌基準自動更新腳本
用法: python update_embedded.py <excel路徑> <html路徑> [--allow-key-changes]

解析規則(與儀表板 JS 解析器一致):
- 只取第 3 列有日期的欄
- 排除第 5 列日號為英文月份縮寫的月均欄(如 MAY、JUL);同日期多欄取先出現者
- 排除週末合併欄(第 3 列無日期,自然排除)與「-」休市值
- 排除全為整數且(≤31 或嚴格遞增)的輔助列
- A 欄品名合併儲存格向下填補;品名 key = 品名 + 空格 + 交易條件
- 數據截止日 = 最後一個有數值的日期欄
安全機制:品名 key 與現有 EMBEDDED 不一致時中止(避免壞檔弄壞網站),
除非加 --allow-key-changes。
"""
import sys, re, json, math, datetime, warnings

def num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if math.isfinite(v) else None
    try:
        return float(str(v).strip().replace(',', ''))
    except ValueError:
        return None

def clean(v):
    if v is None:
        return None
    v = round(v, 6)
    return int(v) if float(v).is_integer() else v

def parse_excel(path):
    import openpyxl
    warnings.filterwarnings('ignore')
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        sys.exit('錯誤:無法開啟 %s(不是有效的 .xlsx?):%s' % (path, e))
    ws = wb[wb.sheetnames[0]]

    month_tag = re.compile(r'^(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)[A-Z]*$', re.I)
    date_cols, seen, excluded = [], set(), []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(3, c).value
        if not isinstance(v, datetime.datetime):
            continue
        d = v.strftime('%Y-%m-%d')
        tag = ws.cell(5, c).value
        if isinstance(tag, str) and month_tag.fullmatch(tag.strip()):
            excluded.append((d, tag.strip())); continue
        if d in seen:
            excluded.append((d, 'dup')); continue
        seen.add(d)
        date_cols.append((c, d))
    if len(date_cols) < 2:
        sys.exit('錯誤:第 3 列找不到足夠的日期欄,請確認 Excel 格式')

    products, series, carry = [], {}, ''
    for r in range(4, ws.max_row + 1):
        name_raw, term_raw = ws.cell(r, 1).value, ws.cell(r, 2).value
        name = re.sub(r'\s*\n\s*', ' ', str(name_raw)).strip() if name_raw is not None else ''
        term = str(term_raw).strip() if term_raw is not None else ''
        if name:
            carry = name
        else:
            name = carry
        vals = [num(ws.cell(r, c).value) for c, d in date_cols]
        vals_nn = [v for v in vals if v is not None]
        if not vals_nn:
            continue
        if all(float(v).is_integer() for v in vals_nn):
            inc = all(vals_nn[i] > vals_nn[i-1] for i in range(1, len(vals_nn)))
            if max(vals_nn) <= 31 or inc:
                continue
        if re.search(r'品名|产品|product|item', name, re.I) and re.search(r'交易|条件|term', term, re.I):
            continue
        full = (name + ' ' + term).strip()
        key, k = full, 2
        while key in series:
            key = full + ' #%d' % k; k += 1
        series[key] = {d: num(ws.cell(r, c).value) for c, d in date_cols
                       if num(ws.cell(r, c).value) is not None}
        products.append(key)

    cutoff = max(d for m in series.values() for d in m)
    dates = [d for c, d in date_cols if d <= cutoff]
    emb = {'dates': dates,
           'series': {p: [clean(series[p].get(d)) for d in dates] for p in products}}
    return emb, cutoff, excluded

def main():
    args = [a for a in sys.argv[1:] if not a.startswith('--')]
    allow_key_changes = '--allow-key-changes' in sys.argv
    if len(args) != 2:
        sys.exit(__doc__)
    xlsx_path, html_path = args

    emb, cutoff, excluded = parse_excel(xlsx_path)

    html = open(html_path, encoding='utf-8').read()
    m = re.search(r'const EMBEDDED = (\{.*?\});\n', html, re.S)
    if not m:
        sys.exit('錯誤:HTML 中找不到 EMBEDDED 常數')
    old = json.loads(m.group(1))

    old_keys, new_keys = list(old['series'].keys()), list(emb['series'].keys())
    if old_keys != new_keys:
        added = [k for k in new_keys if k not in old_keys]
        removed = [k for k in old_keys if k not in new_keys]
        msg = ('品名 key 與現有基準不一致!\n  新增: %s\n  移除: %s' % (added, removed))
        if not allow_key_changes:
            sys.exit('中止(未修改網站):' + msg +
                     '\n若確定要變更品名,加上 --allow-key-changes 重跑。')
        print('警告(--allow-key-changes 已啟用):' + msg)

    emb_str = json.dumps(emb, ensure_ascii=False, separators=(',', ':'))
    new_html, n = re.subn(r'const EMBEDDED = \{.*?\};\n',
                          lambda _: 'const EMBEDDED = ' + emb_str + ';\n',
                          html, count=1, flags=re.S)
    if n != 1:
        sys.exit('錯誤:EMBEDDED 替換失敗')
    open(html_path, 'w', encoding='utf-8').write(new_html)

    old_last = old['dates'][-1]
    print('✔ 更新完成')
    print('  產品數      :', len(new_keys))
    print('  交易日      :', len(emb['dates']), '(%s ～ %s)' % (emb['dates'][0], cutoff))
    print('  基準截止    : %s → %s' % (old_last, cutoff))
    if excluded:
        print('  已排除月均/重複欄:', len(excluded), '個', sorted(set(t for _, t in excluded)))

if __name__ == '__main__':
    main()
